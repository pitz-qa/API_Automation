import openpyxl
import requests
import json
import jsonpath

class Common:

    def __init__(self,pathfile,sheetname):
        global wk
        global sh
        wk = openpyxl.load_workbook(pathfile)
        sh = wk[sheetname]

    def fetch_excel_rows(self):
        rows = sh.max_row
        return rows

    def fetch_col_count(self):
        col = sh.max_column
        return col

    def fetch_key_name(self):
        c =sh.max_column
        li = []
        for i in range(1,c+1):
            cell = sh.cell(row=1,column=i)
            li.insert(i-1,cell.value)

        return li

    def update_request_withData(self,rowNumber,jsonRequest,keyList):
        c = sh.max_column
        for i in range(1,c+1):
            cell = sh.cell(row=rowNumber,column=i)
            jsonRequest[keyList[i-1]]=cell.value

        return jsonRequest
