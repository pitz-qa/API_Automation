# python3 -m pytest testCase_001_ValidateRegister.py -v -s
# source env/bin/activate
# pytest -k register Smoke groupTC.py -v -s (Specific test cases)
# pytest -m Smoke groupTC.py -v -s (run group of TC)
# pytest -m "not Smoke" groupTC.py -v -s (not run of TC)
import openpyxl
import requests
import json

def test_Add_New_StudentData():
    url = "http://www.thetestingworldapi.com/api/studentsDetails"
    f = open('/Users/user/Desktop/APIAutomation/api-testing-python-udemy/TestCases/DataDriven/createStudentData.json', 'r')
    json_request = json.loads(f.read())

    wk = openpyxl.load_workbook('/Users/user/Desktop/APIAutomation/api-testing-python-udemy/TestCases/DataDriven/studentdata.xlsx')
    sh = wk['Sheet1']
    rows = sh.max_row

    for i in range(2,rows+1):
        cell_first_name = sh.cell(row=i, column=1)
        cell_mid_name = sh.cell(row=i, column=2)
        cell_last_name = sh.cell(row=i, column=3)
        cell_dob = sh.cell(row=i, column=4)

        json_request['first_name'] = cell_first_name.value
        json_request['middle_name'] = cell_mid_name.value
        json_request['last_name'] = cell_last_name.value
        json_request['date_of_birth'] = cell_dob.value

        response = requests.post(url, json_request)
        print(response.text)
        print(response.status_code)
        assert response.status_code == 201